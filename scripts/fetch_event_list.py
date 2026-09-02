#!/usr/bin/env python3
"""Fetch a parent-supplied event list from a URL, for >add-event's URL form.

This script's job stops at "turn a URL into structured rows with the
sheet's own column headers" - it does NOT decide which columns matter,
what counts as a valid date, or which rows describe official-vs-fundraising
events. Those are the same judgment calls Claude already makes reading raw
page text elsewhere in this skill, and belong there, not baked into a
parser here.

Supports three URL shapes:
  - A Google Sheets "edit" URL (docs.google.com/spreadsheets/d/<id>/...) -
    tries the sheet's public CSV export first. This only works if the
    sheet is shared "Anyone with the link" (same requirement as the
    calendar .ics files themselves) - if it's not, Google redirects to a
    login page instead of CSV data, which this script detects and reports
    as `requires_drive_tool: true` rather than silently returning garbage.
    When that happens, the caller (Claude, per SKILL.md) should fall back
    to a connected Google Drive tool in the current session - that path
    only exists at the Claude-tooling layer, this script has no way to
    authenticate to Drive on its own.
  - A direct .xlsx URL - downloaded and parsed with openpyxl.
  - A direct .csv URL (or anything else - falls back to treating the
    response body as CSV text, since that's the most common shape for a
    "here's a link to our events list" URL).

Usage:
    python3 fetch_event_list.py --url <url>
"""
from __future__ import annotations

import argparse
import csv
import io
import json
import re

import requests

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)
TIMEOUT = 20

_GOOGLE_SHEETS_RE = re.compile(r"docs\.google\.com/spreadsheets/d/([a-zA-Z0-9_-]+)")
_GID_RE = re.compile(r"[?#&]gid=(\d+)")


def _rows_from_csv_text(text: str) -> list[dict]:
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def _looks_like_html(text: str) -> bool:
    stripped = text.lstrip().lower()
    return stripped.startswith("<!doctype") or stripped.startswith("<html")


def fetch_google_sheet(url: str) -> dict:
    match = _GOOGLE_SHEETS_RE.search(url)
    if not match:
        return {"error": True, "message": "Couldn't find a Google Sheets file id in that URL."}
    sheet_id = match.group(1)
    gid_match = _GID_RE.search(url)
    gid = gid_match.group(1) if gid_match else "0"

    export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"
    try:
        response = requests.get(
            export_url, headers={"User-Agent": USER_AGENT}, timeout=TIMEOUT, allow_redirects=True
        )
    except requests.RequestException as exc:
        return {"error": True, "message": f"Couldn't reach that Google Sheet: {exc}"}

    if response.status_code != 200 or _looks_like_html(response.text):
        return {
            "requires_drive_tool": True,
            "sheet_id": sheet_id,
            "message": (
                "This sheet isn't publicly viewable (or isn't shared 'Anyone with the "
                "link'), so it can't be fetched directly. Use a connected Google Drive "
                "tool in this session instead, e.g. download_file_content with "
                f"fileId={sheet_id} and exportMimeType text/csv."
            ),
        }

    rows = _rows_from_csv_text(response.text)
    return {"source": "google_sheets", "sheet_id": sheet_id, "gid": gid, "rows": rows}


def fetch_xlsx(url: str) -> dict:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"error": True, "message": "openpyxl isn't installed - run check_dependencies.py --install first."}

    try:
        response = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=TIMEOUT)
        response.raise_for_status()
    except requests.RequestException as exc:
        return {"error": True, "message": f"Couldn't download that file: {exc}"}

    try:
        workbook = load_workbook(io.BytesIO(response.content), data_only=True)
    except Exception as exc:
        return {"error": True, "message": f"Couldn't read this as an Excel file: {exc}"}

    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    except StopIteration:
        return {"source": "xlsx", "rows": []}

    rows = []
    for raw_row in rows_iter:
        row = {header[i]: raw_row[i] for i in range(min(len(header), len(raw_row)))}
        if any(v not in (None, "") for v in row.values()):
            rows.append({k: ("" if v is None else str(v)) for k, v in row.items()})

    return {"source": "xlsx", "rows": rows}


def fetch_csv(url: str) -> dict:
    try:
        response = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=TIMEOUT)
        response.raise_for_status()
    except requests.RequestException as exc:
        return {"error": True, "message": f"Couldn't download that file: {exc}"}

    if _looks_like_html(response.text):
        return {"error": True, "message": "That URL returned a webpage, not CSV data - check it's a direct file link."}

    return {"source": "csv", "rows": _rows_from_csv_text(response.text)}


def fetch(url: str) -> dict:
    if "docs.google.com/spreadsheets" in url:
        return fetch_google_sheet(url)
    path = url.split("?")[0].lower()
    if path.endswith(".xlsx"):
        return fetch_xlsx(url)
    return fetch_csv(url)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--url", required=True)
    args = parser.parse_args()
    print(json.dumps(fetch(args.url), indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
